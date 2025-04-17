import sys
import json
import os
import webbrowser
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QFileDialog, QLabel, QTextEdit, QComboBox, QGridLayout, QLineEdit, QMessageBox, QStyleFactory
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtCore import Qt, pyqtSlot, pyqtSignal, QUrl, QTimer, QObject
from PyQt5.QtWebEngineWidgets import QWebEngineView
from openpyxl import load_workbook
import traceback
from PyQt5.QtWebChannel import QWebChannel

class Bridge(QObject):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window

    @pyqtSlot(str)
    def receiveStatus(self, status):
        print("[Web Scraping Debug] Received status:", status)  # Debug print
        # Update the browser window's status label for user feedback
        if hasattr(self.main_window, 'browser_status_label'):
            self.main_window.browser_status_label.setText(f"Web Scraping Status: {status}")
        self.main_window.status_label.setText(status)
        
        # Handle fallback case
        if status == 'IN_PROGRESS_FALLBACK':
            self.main_window.build_state_dropdown.setCurrentText("In Progress")
            self.main_window.status_label.setText("⚠️ Design tab not found – set to In Progress")
            if hasattr(self.main_window, 'timer'):
                self.main_window.timer.stop()
            return
            
        # If design is approved, update the Excel file
        if "✅ Design Approved" in status:
            try:
                # Extract PID and Node from status message
                pid_match = status.split("PID:")[1].split("|")[0].strip()
                node_match = status.split("Node:")[1].strip()
                
                # Find matching PID input field or first empty one
                pid_updated = False
                for i, pid_input in enumerate(self.main_window.pid_inputs):
                    if pid_input.text().strip() == pid_match:
                        # Update the corresponding node input
                        self.main_window.node_inputs[i].setText(node_match)
                        # Set build state to "Design Approved"
                        self.main_window.build_state_dropdown.setCurrentText("Design Approved")
                        pid_updated = True
                        break
                
                # If no match found, fill the first empty PID field
                if not pid_updated:
                    for i, pid_input in enumerate(self.main_window.pid_inputs):
                        if not pid_input.text().strip():
                            pid_input.setText(pid_match)
                            self.main_window.node_inputs[i].setText(node_match)
                            self.main_window.build_state_dropdown.setCurrentText("Design Approved")
                            pid_updated = True
                            break
                
                if pid_updated:
                    # Update Excel with approval status
                    self.main_window.update_excel(pid_match, node_match)
                    # Stop the timer since we found a successful status
                    if hasattr(self.main_window, 'timer'):
                        self.main_window.timer.stop()
                
            except Exception as e:
                print(f"Error updating Excel: {str(e)}")
                traceback.print_exc()

class ExcelAutomationApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Automation App")
        self.setGeometry(100, 100, 800, 500)  # Increased window size
        
        # Initialize WebChannel for JavaScript-Python communication
        self.channel = QWebChannel()
        
        # Set up theme
        self.setup_theme()

        # Layout
        layout = QGridLayout()

        # Top row - Browser controls
        browser_label = QLabel("Browser:")
        layout.addWidget(browser_label, 0, 0)
        
        self.browser_dropdown = QComboBox()
        self.browser_dropdown.addItems(["Edge", "Chrome"])
        layout.addWidget(self.browser_dropdown, 0, 1)

        self.open_browser_button = QPushButton("Open Browser")
        self.open_browser_button.clicked.connect(self.open_browser)
        layout.addWidget(self.open_browser_button, 0, 2)

        # Second row - Excel controls
        self.load_button = QPushButton("Load Excel File")
        self.load_button.clicked.connect(self.load_excel)
        layout.addWidget(self.load_button, 1, 0, 1, 3)  # Span 3 columns

        # PID Inputs - Start from row 2
        self.pid_inputs = [QLineEdit() for _ in range(4)]
        for i, pid_input in enumerate(self.pid_inputs):
            pid_input.setPlaceholderText(f"PID {i + 1}")
            layout.addWidget(pid_input, 2, i)

        # Node Inputs
        self.node_inputs = [QLineEdit() for _ in range(4)]
        for i, node_input in enumerate(self.node_inputs):
            node_input.setPlaceholderText(f"Node {i + 1}")
            layout.addWidget(node_input, 3, i)

        # Scope Inputs
        self.scope_inputs = [QLineEdit() for _ in range(4)]
        for i, scope_input in enumerate(self.scope_inputs):
            scope_input.setPlaceholderText(f"Scope {i + 1}")
            layout.addWidget(scope_input, 4, i)

        # Magellan Inputs
        self.magellan_inputs = [QLineEdit() for _ in range(4)]
        for i, magellan_input in enumerate(self.magellan_inputs):
            magellan_input.setPlaceholderText(f"Magellan {i + 1}")
            layout.addWidget(magellan_input, 5, i)

        # Config and Build State row (use QHBoxLayout for perfect alignment)
        config_build_row = QHBoxLayout()
        self.config_dropdown = QComboBox()
        self.config_dropdown.addItems(["1x1", "2x2", "4x4", "N/A"])
        self.config_dropdown.setMinimumWidth(180)
        self.config_dropdown.setMaximumWidth(180)
        config_build_row.addWidget(self.config_dropdown)

        self.build_state_dropdown = QComboBox()
        self.build_state_dropdown.setEditable(True)
        self.build_state_dropdown.addItems(["In Design", "In Progress", "Does Not Exist", "PRO-I", "Design Approved"])
        self.build_state_dropdown.setMinimumWidth(180)
        self.build_state_dropdown.setMaximumWidth(180)
        config_build_row.addWidget(self.build_state_dropdown)

        layout.addLayout(config_build_row, 6, 0, 1, 4)

        # Action buttons row
        self.save_next_button = QPushButton("Save & Next")
        self.save_next_button.clicked.connect(self.save_next_action)
        layout.addWidget(self.save_next_button, 7, 0)

        self.back_button = QPushButton("Back")
        self.back_button.clicked.connect(self.load_previous_row)
        layout.addWidget(self.back_button, 7, 1)

        self.toggle_dark_mode_button = QPushButton("Toggle Dark Mode")
        self.toggle_dark_mode_button.clicked.connect(self.toggle_dark_mode)
        layout.addWidget(self.toggle_dark_mode_button, 7, 2, 1, 2)  # Span 2 columns

        # Status labels
        self.status_label = QLabel("No file loaded.")
        layout.addWidget(self.status_label, 8, 0, 1, 4)  # Span all columns

        self.last_node_label = QLabel("Last Node: N/A")
        layout.addWidget(self.last_node_label, 9, 0, 1, 4)  # Span all columns

        # Row navigation
        self.row_input = QLineEdit()
        self.row_input.setPlaceholderText("Enter row #")
        layout.addWidget(self.row_input, 10, 0)

        self.go_button = QPushButton("Go to Row")
        self.go_button.clicked.connect(self.load_specific_row)
        layout.addWidget(self.go_button, 10, 1)

        self.open_excel_button = QPushButton("Open in Excel (Read-Only)")
        self.open_excel_button.clicked.connect(self.open_excel_readonly)
        layout.addWidget(self.open_excel_button, 10, 2, 1, 2)  # Span 2 columns

        self.current_row_label = QLabel("Current Row: N/A")
        layout.addWidget(self.current_row_label, 11, 0, 1, 4)  # Span all columns

        # Main Widget
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        # Placeholder for Excel file path
        self.excel_file_path = None

    def setup_theme(self, dark_mode=False):
        app = QApplication.instance()
        palette = QPalette()

        if dark_mode:
            # Dark mode settings
            palette.setColor(QPalette.Window, QColor(53, 53, 53))
            palette.setColor(QPalette.WindowText, Qt.white)
            palette.setColor(QPalette.Base, QColor(35, 35, 35))
            palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
            palette.setColor(QPalette.ToolTipBase, QColor(25, 25, 25))
            palette.setColor(QPalette.ToolTipText, Qt.white)
            palette.setColor(QPalette.Text, Qt.white)
            palette.setColor(QPalette.Button, QColor(53, 53, 53))
            palette.setColor(QPalette.ButtonText, Qt.white)
            palette.setColor(QPalette.BrightText, Qt.red)
            palette.setColor(QPalette.Link, QColor(42, 130, 218))
            palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
            palette.setColor(QPalette.HighlightedText, QColor(35, 35, 35))
            
            # Set the dark style sheet for QComboBox and QLineEdit
            self.setStyleSheet("""
                QComboBox {
                    background-color: #353535;
                    color: white;
                    border: 1px solid #555555;
                    padding: 5px;
                }
                QComboBox:drop-down {
                    border: 1px solid #555555;
                }
                QComboBox:down-arrow {
                    width: 15px;
                    height: 15px;
                }
                QLineEdit {
                    background-color: #353535;
                    color: white;
                    border: 1px solid #555555;
                    padding: 5px;
                }
                QPushButton {
                    background-color: #454545;
                    color: white;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                }
                QPushButton:hover {
                    background-color: #555555;
                }
                QPushButton:pressed {
                    background-color: #252525;
                }
            """)
        else:
            # Light mode settings
            palette = app.style().standardPalette()
            self.setStyleSheet("")

        app.setPalette(palette)

    def show_dark_messagebox(self, icon, title, text):
        msg = QMessageBox(self)
        msg.setIcon(icon)
        msg.setWindowTitle(title)
        msg.setText(text)
        if QApplication.instance().palette().color(QPalette.Window).lightness() < 128:
            # Dark mode: set dark palette and style
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: #353535;
                    color: white;
                }
                QPushButton {
                    background-color: #454545;
                    color: white;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                }
                QPushButton:hover {
                    background-color: #555555;
                }
                QPushButton:pressed {
                    background-color: #252525;
                }
            """)
        msg.exec_()

    def load_excel(self):
        # Open file dialog to select Excel file
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            try:
                # Try to load the workbook to check if it's valid
                workbook = load_workbook(file_path, read_only=True)
                workbook.close()
                
                self.excel_file_path = file_path
                self.status_label.setText(f"Loaded: {file_path}")
                self.show_dark_messagebox(QMessageBox.Information, "Excel Loaded", f"Loaded: {os.path.basename(file_path)}")
            except Exception as e:
                self.show_dark_messagebox(QMessageBox.Critical, "Error", f"Failed to load Excel file: {str(e)}")

    def save_next_action(self):
        if not self.excel_file_path:
            self.status_label.setText("No Excel file loaded.")
            return

        try:
            workbook = load_workbook(self.excel_file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            # Decide which row to write to
            row = self.current_row if self.current_row else sheet.max_row + 1

            # Write PIDs
            for i, pid_input in enumerate(self.pid_inputs):
                col_name = f"PID {i+1}"
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    sheet.cell(row=row, column=col_idx).value = pid_input.text().strip().replace('\u00A0', '').replace('\u200B', '')

            # Write Scope - Fix spacing issue in header name
            for i, scope_input in enumerate(self.scope_inputs):
                # Try both with single and double space
                col_name1 = f"SCOPE {i+1}"
                col_name2 = f"SCOPE  {i+1}"
                if col_name1 in headers:
                    col_idx = headers.index(col_name1) + 1
                    sheet.cell(row=row, column=col_idx).value = scope_input.text().strip().replace('\u00A0', '').replace('\u200B', '')
                elif col_name2 in headers:
                    col_idx = headers.index(col_name2) + 1
                    sheet.cell(row=row, column=col_idx).value = scope_input.text().strip().replace('\u00A0', '').replace('\u200B', '')

            # Write Magellan - Fix spacing issue in header name
            for i, mag_input in enumerate(self.magellan_inputs):
                col_name1 = f"MAGELLAN {i+1}"
                col_name2 = f"MAGELLAN  {i+1}"
                if col_name1 in headers:
                    col_idx = headers.index(col_name1) + 1
                    sheet.cell(row=row, column=col_idx).value = mag_input.text().strip().replace('\u00A0', '').replace('\u200B', '')
                elif col_name2 in headers:
                    col_idx = headers.index(col_name2) + 1
                    sheet.cell(row=row, column=col_idx).value = mag_input.text().strip().replace('\u00A0', '').replace('\u200B', '')

            # Write NODE
            for i, node_input in enumerate(self.node_inputs):
                col_name = f"NODE {i+1}"
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    sheet.cell(row=row, column=col_idx).value = node_input.text().strip().replace('\u00A0', '').replace('\u200B', '')

            # Try multiple possible column names for AOI NODE or CONFIG
            config_value = self.config_dropdown.currentText().strip().replace('\u00A0', '').replace('\u200B', '')
            for col_name in ["AOI NODE", "CONFIG", "NODE CONFIG"]:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    sheet.cell(row=row, column=col_idx).value = config_value
                    break

            # Try multiple possible column names for NOTES or BUILD STATE
            build_state_value = self.build_state_dropdown.currentText().strip().replace('\u00A0', '').replace('\u200B', '')
            for col_name in ["NOTES", "BUILD STATE", "STATE"]:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    sheet.cell(row=row, column=col_idx).value = build_state_value
                    break

            # Update labels
            self.last_node_label.setText(f"Last Node: {self.magellan_inputs[0].text()}")
            self.current_row_label.setText(f"Current Row: {row}")
            self.status_label.setText(f"Saved row {row}")

            workbook.save(self.excel_file_path)
            self.show_dark_messagebox(QMessageBox.Information, "Saved", f"Row {row} saved successfully!")

            # Clear inputs
            for input_list in [self.pid_inputs, self.scope_inputs, self.magellan_inputs, self.node_inputs]:
                for field in input_list:
                    field.clear()

            self.config_dropdown.setCurrentIndex(0)
            self.build_state_dropdown.setCurrentIndex(0)

            # Move to next row
            self.current_row = row + 1
            
        except Exception as e:
            self.show_dark_messagebox(QMessageBox.Critical, "Error", f"Failed to save data: {str(e)}")
            traceback.print_exc()

    def load_previous_row(self):
        if not self.excel_file_path:
            self.status_label.setText("No Excel file loaded.")
            return

        try:
            workbook = load_workbook(self.excel_file_path)
            sheet = workbook.active

            if self.current_row is None:
                self.current_row = sheet.max_row
            else:
                self.current_row = max(2, self.current_row - 1)

            headers = [cell.value for cell in sheet[1]]
            self.load_row_data(sheet, headers)
            
            self.status_label.setText(f"Loaded previous row: {self.current_row}")
            self.current_row_label.setText(f"Current Row: {self.current_row}")
            
        except Exception as e:
            self.show_dark_messagebox(QMessageBox.Critical, "Error", f"Failed to load previous row: {str(e)}")

    def load_specific_row(self):
        if not self.excel_file_path:
            self.status_label.setText("No Excel file loaded.")
            return

        try:
            target_row = int(self.row_input.text().strip())
        except ValueError:
            self.status_label.setText("Invalid row number.")
            return

        try:
            workbook = load_workbook(self.excel_file_path)
            sheet = workbook.active

            if target_row == 1:
                self.show_dark_messagebox(QMessageBox.Information, "Invalid Row", "Row 1 contains headers and cannot be edited.")
                self.status_label.setText("Row 1 contains headers and cannot be edited.")
                return
            elif target_row < 2 or target_row > sheet.max_row:
                self.status_label.setText("Row number out of range.")
                return

            self.current_row = target_row
            headers = [cell.value for cell in sheet[1]]
            
            self.load_row_data(sheet, headers)
            
            self.status_label.setText(f"Loaded row: {self.current_row}")
            self.current_row_label.setText(f"Current Row: {self.current_row}")
            
        except Exception as e:
            self.show_dark_messagebox(QMessageBox.Critical, "Error", f"Failed to load row {target_row}: {str(e)}")

    def load_row_data(self, sheet, headers):
        """Helper method to load data from the current row into the input fields"""
        # Load values back into input fields
        for i in range(4):
            # Get PID values
            for col_name in [f"PID {i+1}"]:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    self.pid_inputs[i].setText(str(sheet.cell(row=self.current_row, column=col_idx).value or ""))

            # Get Scope values - try both spacing variants
            for col_name in [f"SCOPE {i+1}", f"SCOPE  {i+1}"]:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    self.scope_inputs[i].setText(str(sheet.cell(row=self.current_row, column=col_idx).value or ""))

            # Get Magellan values - try both spacing variants
            for col_name in [f"MAGELLAN {i+1}", f"MAGELLAN  {i+1}"]:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    self.magellan_inputs[i].setText(str(sheet.cell(row=self.current_row, column=col_idx).value or ""))

            # Get Node values
            for col_name in [f"NODE {i+1}"]:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    self.node_inputs[i].setText(str(sheet.cell(row=self.current_row, column=col_idx).value or ""))

        # Try multiple column names for CONFIG
        config_found = False
        for config_col_name in ["CONFIG", "AOI NODE", "NODE CONFIG"]:
            if config_col_name in headers:
                col_idx = headers.index(config_col_name) + 1
                config_value = sheet.cell(row=self.current_row, column=col_idx).value
                if config_value in [self.config_dropdown.itemText(i) for i in range(self.config_dropdown.count())]:
                    self.config_dropdown.setCurrentText(str(config_value))
                    config_found = True
                    break

        # Try multiple column names for BUILD STATE
        state_found = False
        for state_col_name in ["BUILD STATE", "NOTES", "STATE"]:
            if state_col_name in headers:
                col_idx = headers.index(state_col_name) + 1
                state_value = sheet.cell(row=self.current_row, column=col_idx).value
                if state_value:
                    self.build_state_dropdown.setCurrentText(str(state_value))
                    state_found = True
                    break

        # Update node label
        self.last_node_label.setText(f"Last Node: {self.magellan_inputs[0].text()}")

    def open_excel_readonly(self):
        if not self.excel_file_path:
            self.status_label.setText("No Excel file loaded.")
            return

        try:
            if sys.platform == "win32":
                # Use the /r switch with start to explicitly open Excel in read-only mode
                os.system(f'start "" "EXCEL.EXE" /r "{self.excel_file_path}"')
            elif sys.platform == "darwin":
                os.system(f'open -a "Microsoft Excel" "{self.excel_file_path}"')
            else:
                os.system(f'libreoffice --view "{self.excel_file_path}"')

            self.status_label.setText("Opened Excel in read-only mode.")
        except Exception as e:
            self.status_label.setText(f"Failed to open Excel: {e}")
            self.show_dark_messagebox(QMessageBox.Critical, "Error", f"Failed to open Excel: {str(e)}")

    def toggle_dark_mode(self):
        current_palette = QApplication.instance().palette()
        is_dark_mode = current_palette.color(QPalette.Window).lightness() < 128
        self.setup_theme(not is_dark_mode)  # Toggle the theme

    def open_browser(self):
        try:
            self.browser_window = QMainWindow()
            self.browser_window.setWindowTitle("Design Review")
            self.browser_window.resize(1200, 800)
            self.web_view = QWebEngineView()
            main_widget = QWidget()
            layout = QVBoxLayout()
            main_widget.setLayout(layout)
            self.browser_status_label = QLabel("Web Scraping Status: (waiting)")
            self.browser_status_label.setStyleSheet("background: #222; color: #fff; padding: 4px; font-weight: bold;")
            layout.addWidget(self.browser_status_label)
            nav_bar = QHBoxLayout()
            self.url_input = QLineEdit()
            self.url_input.setPlaceholderText("Enter URL")
            self.url_input.setText("https://www.google.com")
            go_button = QPushButton("Go")
            go_button.clicked.connect(self.navigate_to_url)
            back_button = QPushButton("◀")
            back_button.clicked.connect(lambda: self.web_view.back())
            forward_button = QPushButton("▶")
            forward_button.clicked.connect(lambda: self.web_view.forward())
            reload_button = QPushButton("⟳")
            reload_button.clicked.connect(lambda: self.web_view.reload())
            home_button = QPushButton("🏠")
            home_button.clicked.connect(lambda: self.web_view.setUrl(QUrl("https://www.google.com")))
            nav_bar.addWidget(back_button)
            nav_bar.addWidget(forward_button)
            nav_bar.addWidget(reload_button)
            nav_bar.addWidget(home_button)
            nav_bar.addWidget(self.url_input)
            nav_bar.addWidget(go_button)
            layout.addLayout(nav_bar)
            layout.addWidget(self.web_view)
            scrape_button = QPushButton("Scrape Now")
            scrape_button.clicked.connect(self.check_status)
            layout.addWidget(scrape_button)
            # Add Show Page Source button
            show_source_button = QPushButton("Show Page Source")
            def show_source():
                self.web_view.page().toHtml(lambda html: self.show_dark_messagebox(QMessageBox.Information, "Page Source", html[:2000] + ("\n...truncated..." if len(html) > 2000 else "")))
            show_source_button.clicked.connect(show_source)
            layout.addWidget(show_source_button)
            self.browser_window.setCentralWidget(main_widget)
            self.bridge = Bridge(self)
            self.web_view.page().setWebChannel(self.channel)
            self.channel.registerObject('bridge', self.bridge)
            # Inject the Qt WebChannel JS so window.bridge is available
            self.web_view.page().runJavaScript('''
                if (!window.qt) {
                    var s = document.createElement('script');
                    s.src = 'qrc:///qtwebchannel/qwebchannel.js';
                    s.onload = function() {
                        new QWebChannel(qt.webChannelTransport, function(channel) {
                            window.bridge = channel.objects.bridge;
                        });
                    };
                    document.head.appendChild(s);
                } else {
                    new QWebChannel(qt.webChannelTransport, function(channel) {
                        window.bridge = channel.objects.bridge;
                    });
                }
            ''')
            prism_url = "https://www.google.com"
            self.web_view.setUrl(QUrl(prism_url))
            self.web_view.loadFinished.connect(self.inject_dark_css_if_needed)
            self.timer = QTimer()
            self.timer.timeout.connect(self.check_status)
            self.timer.start(5000)
            self.browser_window.show()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open browser: {str(e)}")
            print(f"Error opening browser: {str(e)}")
            traceback.print_exc()

    def inject_dark_css_if_needed(self):
        if QApplication.instance().palette().color(QPalette.Window).lightness() < 128:
            dark_css = """
                html, body {
                    background: #232323 !important;
                    color: #eee !important;
                }
                table, td, th {
                    background: #232323 !important;
                    color: #eee !important;
                    border-color: #444 !important;
                }
                input, textarea, select {
                    background: #353535 !important;
                    color: #fff !important;
                    border: 1px solid #555 !important;
                }
                a { color: #8ecfff !important; }
            """
            js = f"""
                var style = document.getElementById('darkmode-style');
                if (!style) {{
                    style = document.createElement('style');
                    style.id = 'darkmode-style';
                    style.innerHTML = `{dark_css}`;
                    document.head.appendChild(style);
                }}
            """
            self.web_view.page().runJavaScript(js)

    def navigate_to_url(self):
        url = self.url_input.text().strip()
        if not url.startswith("http"):
            url = "https://" + url
        self.web_view.setUrl(QUrl(url))

    def check_status(self):
        print("check_status called")  # Debug print
        js_code = """
        (function() {
            function extractStatus() {
                const labels = document.querySelectorAll('td.formLabel');
                let pid = '';
                let node = '';
                let status = '';
                for (const label of labels) {
                    const labelText = label.textContent.trim();
                    const valueCell = label.nextElementSibling;
                    const value = valueCell ? valueCell.textContent.trim() : '';
                    if (labelText.includes('PID')) {
                        pid = value;
                    } else if (labelText.includes('Node')) {
                        node = value;
                    } else if (labelText.includes('Design Status')) {
                        status = value;
                    }
                }
                if (pid && node) {
                    if (status.toLowerCase().includes('design approved')) {
                        return `✅ Design Approved | PID: ${pid} | Node: ${node}`;
                    } else if (status.toLowerCase().includes('in progress')) {
                        return `⏳ Design In Progress | PID: ${pid} | Node: ${node}`;
                    } else if (status.toLowerCase().includes('rejected')) {
                        return `❌ Design Rejected | PID: ${pid} | Node: ${node}`;
                    }
                }
                return 'IN_PROGRESS_FALLBACK';
            }
            console.log('JS running, window.bridge is', typeof window.bridge);
            if (typeof window.bridge !== 'undefined') {
                window.bridge.receiveStatus(extractStatus());
            }
        })();
        """
        self.web_view.page().runJavaScript(js_code)

# Run the app
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelAutomationApp()
    window.show()
    sys.exit(app.exec_())